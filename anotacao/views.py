from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.db.models import Q
from django.contrib.auth.hashers import make_password, check_password
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.core.files.base import ContentFile
import io

import pandas as pd
import openpyxl

from anotacao.models import UserMilitar, Anotacao, TurmaCadetes, ArquivoExportado
from anotacao.forms import LoginForm, AlterarSenhaForm,AnotacaoForm
from anotacao.decorators import usuario_autenticado

def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            nome_ou_numerica = form.cleaned_data['nome_ou_numerica']
            senha = form.cleaned_data['senha']

            usuario = UserMilitar.objects.filter(
                Q(nome=nome_ou_numerica) | Q(numerica=nome_ou_numerica)
            ).first()

            if usuario: 
                if senha == usuario.senha or check_password(senha, usuario.senha):
                    request.session['user_id'] = usuario.id
                    return redirect('home')
                else:
                    messages.error(request, 'Senha Incorreta')
            else:
                messages.error(request, 'Usuário não encontrado')

    else:
        form = LoginForm()

    return render(request, 'anotacao/login.html', {'form': form})

def logout_view(request):
    request.session.flush()
    return redirect('login')

@usuario_autenticado
def alterar_senha_view(request):
    usuario = request.usuario

    if request.method == 'POST':
        form = AlterarSenhaForm(request.POST, usuario=usuario)
        if form.is_valid():
            # Se o formulário for válido, atualize a senha do usuário
            nova_senha = form.cleaned_data['nova_senha']
            usuario.senha = make_password(nova_senha)  # Aplica a hash da nova senha
            usuario.save()
            messages.success(request, "Sua senha foi alterada com sucesso!")  # Mensagem de sucesso
            return redirect('home')  # Redireciona para a página inicial
        else:
            # Se o formulário for inválido, exibe os erros
            messages.error(request, "Houve um erro ao alterar sua senha. Verifique os campos abaixo.")
    else:
        form = AlterarSenhaForm(usuario=usuario)

    return render(request, 'anotacao/alterar_senha.html', {
        'usuario': usuario,
        'form': form,
    })

@usuario_autenticado
def home_view(request):
    usuario = request.usuario
    
    return render(request, 'anotacao/home.html', {
        'usuario': usuario,
    })

@usuario_autenticado
def cadete_anotacoes_view(request):
    if request.usuario.cargo != 'Cadete':
        return redirect('home')

    cadete_anotado = request.usuario
    anotacoes_notificadas = Anotacao.objects.filter(cadete_anotado=cadete_anotado, status='notificada')
    anotacoes_despachadas = Anotacao.objects.filter(cadete_anotado=cadete_anotado, status='despachada')

    return render(request, 'anotacao/cadete_anotacoes.html', {
        'usuario': cadete_anotado,
        'anotacoes_notificadas': anotacoes_notificadas,
        'anotacoes_despachadas': anotacoes_despachadas,
    })

@usuario_autenticado
def lista_cadetes_view(request):
    usuario = request.usuario  

    if usuario.cargo not in ['Oficial', 'Oficial_comandante'] and not usuario.permissao_anotar:
        return redirect('home')
    
    turmas = TurmaCadetes.objects.all()
    turma_selecionada = request.GET.get('turma')
    if turma_selecionada:
        cadetes = UserMilitar.objects.filter(cargo='Cadete', turma__id=turma_selecionada).order_by('numerica')
    else:
        cadetes = UserMilitar.objects.filter(cargo='Cadete')

    return render(request, 'anotacao/relacao_cadetes.html', {
        'usuario': usuario,
        'cadetes': cadetes,
        'turmas': turmas,
        'turma_selecionada': turma_selecionada
    })

@usuario_autenticado
def criar_anotacao_ajax_view(request):
    cadete_id = request.GET.get('cadete_id')
    cadete_anotado = get_object_or_404(UserMilitar, id=cadete_id, cargo='Cadete')
    
    usuario = request.usuario  
    if usuario.cargo not in ['Oficial', 'Oficial_comandante'] and not usuario.permissao_anotar:
        return redirect('home')
    
    if request.method == 'POST':
        form = AnotacaoForm(request.POST)
        if form.is_valid():
            anotacao = form.save(commit=False)
            anotacao.cadete_anotado = cadete_anotado 
            anotacao.criado_por = request.usuario  
            anotacao.status = 'notificada'  
            anotacao.criada_em = timezone.now()  
            anotacao.save()
            return JsonResponse({'success': True})  # Retorno de sucesso
    else:
        form = AnotacaoForm()

    # Renderiza o formulário como string HTML e envia via JsonResponse
    html_form = render_to_string('anotacao/criar_anotacao_ajax.html', {'form': form, 'cadete_anotado': cadete_anotado, 'usuario': usuario}, request=request)
    return JsonResponse({'html_form': html_form})

@usuario_autenticado
def comandante_anotacoes_view(request):
    usuario = request.usuario
    if usuario.cargo != 'Oficial_comandante':
        return redirect('home')

    turma_selecionada = request.GET.get('turma', '')

    # Filtrar cadetes pela turma e ordenar por 'numerica'
    cadetes = UserMilitar.objects.filter(cargo='Cadete')
    if turma_selecionada:
        cadetes = cadetes.filter(turma_id=turma_selecionada)
    cadetes = cadetes.order_by('numerica')  # Ordena cadetes pela numérica

    # Coletar as anotações para os cadetes, somente as que não foram despachadas
    anotacoes = []
    for cadete in cadetes:
        anotacoes_cadete = Anotacao.objects.filter(
            cadete_anotado=cadete, status='notificada'
        ).order_by('-tipo_anotacao', 'criada_em')  # Ordena por tipo (positivo primeiro) e data
        anotacoes.extend(anotacoes_cadete)

    turmas = TurmaCadetes.objects.all()

    context = {
        'usuario': usuario,
        'anotacoes': anotacoes,
        'turmas': turmas,
        'turma_selecionada': turma_selecionada,
    }
    return render(request, 'anotacao/comandante_anotacoes.html', context)

@usuario_autenticado
def excluir_anotacao_view(request, anotacao_id):
    if request.method == 'POST':
        anotacao = get_object_or_404(Anotacao, id=anotacao_id)
        anotacao.delete()  # Excluir a anotação
        return JsonResponse({'success': True})  # Retorna sucesso

    return JsonResponse({'success': False})  # Se não for um método DELETE ou algo der errado


@usuario_autenticado
def despachar_anotacao_ajax_view(request, anotacao_id):
    if request.usuario.cargo != 'Oficial_comandante':
        return redirect('home')

    anotacao = get_object_or_404(Anotacao, id=anotacao_id)

    if request.method == 'POST':
        if anotacao.tipo_anotacao == 'negativa':
            despacho = request.POST.get('despacho')
            data_cumprimento = request.POST.get('data_cumprimento')
            
            # Salvar a anotação se o despacho for válido
            if despacho in ['pernoite', 'revista', 'outros'] and data_cumprimento:
                if despacho == 'outros':
                    despacho = request.POST.get('outroDespacho')
                    
                anotacao.status = 'despachada'
                anotacao.despacho = despacho
                anotacao.data_cumprimento = data_cumprimento
                anotacao.despachado_por = request.usuario
                anotacao.save()  # Salva a anotação no banco
                return JsonResponse({'success': True})  # Retorna sucesso


        elif anotacao.tipo_anotacao == 'positiva':
            anotacao_negativa_id = request.POST.get('anotacao_negativa_id')
            if anotacao_negativa_id:
                try:
                    anotacao_negativa = Anotacao.objects.get(
                        id=anotacao_negativa_id,
                        cadete_anotado=anotacao.cadete_anotado,
                        tipo_anotacao='negativa',
                        status='notificada'
                    )
                    anotacao_negativa.delete()
                except Anotacao.DoesNotExist:
                    pass
            anotacao.delete()
            return JsonResponse({'success': True})  # Sucesso

    # Carregar as anotações negativas apenas se a anotação for positiva
    anotacoes_negativas = []
    if anotacao.tipo_anotacao == 'positiva':
        anotacoes_negativas = Anotacao.objects.filter(
            cadete_anotado=anotacao.cadete_anotado,
            tipo_anotacao='negativa',
            status='notificada'
        )

    # Renderizar o modal em formato JSON
    html_form = render_to_string(
        'anotacao/despachar_anotacao_ajax.html',
        {'anotacao': anotacao, 'anotacoes_negativas': anotacoes_negativas},
        request=request
    )
    return JsonResponse({'html_form': html_form})


@usuario_autenticado
def anotacoes_despachadas_view(request):
    usuario = request.usuario
    if usuario.cargo != 'Oficial_comandante':
        return redirect('home')

    # Obter anotações despachadas
    anotacoes_despachadas = Anotacao.objects.filter(status='despachada').order_by('cadete_anotado__turma', 'cadete_anotado__numerica')

    # Se o comandante submeter o formulário para exportar
    if request.method == 'POST':
        nome_arquivo = request.POST.get('nome_arquivo')
        if not nome_arquivo:
            nome_arquivo = "anotacoes_despachadas"

        # Preparar os dados para exportação
        dados = []
        for anotacao in anotacoes_despachadas:
            dados.append({
                'Turma': anotacao.cadete_anotado.turma.turma,
                'Cadete': f"{anotacao.cadete_anotado.numerica} - {anotacao.cadete_anotado.nome}",
                'Motivo': f"Anotado Negativamente por {anotacao.criado_por.cargo} {anotacao.criado_por.nome} em {anotacao.criada_em.strftime('%d/%m/%Y %H:%M')} por motivo de: {anotacao.motivo}",
                'Tipo Despacho': f"{anotacao.despacho} - {anotacao.data_cumprimento.strftime('%d/%m/%Y')}",
                'Ass. Ciência': '',
                'Data Cumprimento': '__/__/____',
                'Ass. Cumprimento': ''
            })

        # Criar DataFrame
        df = pd.DataFrame(dados)

        # Criar arquivo Excel em memória
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Anotacoes Despachadas')
            worksheet = writer.sheets['Anotacoes Despachadas']

            # Definindo o tamanho das colunas e formatação
            column_widths = {
                'A': 15,
                'B': 20,
                'C': 50,
                'D': 30,
                'E': 20,
                'F': 20,
                'G': 20,
            }
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width

            # Estilos para cabeçalhos
            header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            header_fill = openpyxl.styles.PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
            header_alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Estilo para células de dados
            thin_border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                 right=openpyxl.styles.Side(style='thin'),
                                                 top=openpyxl.styles.Side(style='thin'),
                                                 bottom=openpyxl.styles.Side(style='thin'))
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='top', wrap_text=True)

        buffer.seek(0)
        excel_file = ContentFile(buffer.read())

        # Criar e salvar o arquivo no modelo ArquivoExportado
        arquivo_exportado = ArquivoExportado.objects.create(
            nome_arquivo=nome_arquivo + '.xlsx',
            exportado_por=request.usuario
        )
        arquivo_exportado.arquivo.save(nome_arquivo + '.xlsx', excel_file)

        # Deletar as anotações despachadas após exportação
        anotacoes_despachadas.delete()

        # Adicionar uma mensagem de sucesso para o usuário
        messages.success(request, 'Anotações exportadas com sucesso para Excel!')

        # Redirecionar de volta para a mesma página
        return redirect('anotacoes_despachadas')

    # Renderizar a página de visualização com o formulário de exportação
    
    return render(request, 'anotacao/anotacoes_despachadas.html', {
        'anotacoes_despachadas': anotacoes_despachadas,
        'usuario': usuario,
    })

@usuario_autenticado
def historico_exportacoes_view(request):
    usuario = request.usuario
    if usuario.cargo != 'Oficial_comandante':
        return redirect('home')

    # Obter os arquivos exportados pelo usuário
    arquivos_exportados = ArquivoExportado.objects.filter(exportado_por=request.usuario).order_by('-data_exportacao')
    
    if request.method == 'POST' and 'limpar_historico' in request.POST:
        arquivos_exportados.delete()
        return redirect('historico_exportacoes')  # Redireciona de volta à página após limpar o histórico
    
    return render(request, 'anotacao/historico_exportacoes.html', {
        'arquivos_exportados': arquivos_exportados,
        'usuario': usuario,
    })
